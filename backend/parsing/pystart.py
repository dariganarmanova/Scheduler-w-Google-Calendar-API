from openpyxl import load_workbook

# Load the workbook and select the active sheet
workbook = load_workbook('export.xlsx')
sheet = workbook.active

# List to store tuples of selected column values from each row
selected_rows = []

# Loop through all rows
for col in sheet.iter_rows(values_only=True):
    selected_tuple = (col[2], col[3], col[4], col[7]) 
    selected_rows.append(selected_tuple) 

for row in selected_rows:
    print(row)

